"""
BSD 3-clause "New" or "Revised" License

Copyright (c) 2023-2032, Arun Bosco J (arun.bosco@samsung.com)
Copyright (c) 2023-2032, Mathi Yugandhararao (yugandhar.m@samsung.com)
All rights reserved.

Redistribution and use in source and binary forms, with or
without modification, are permitted provided that the following
conditions are met:

	* Redistributions of source code must retain the above copyright
	notice, this list of conditions and the following disclaimer.

	* Redistributions in binary form must reproduce the above
	copyright notice, this list of conditions and the following
	disclaimer in the documentation and/or other materials provided
	with the distribution.

	* Neither the name of the copyright holder nor the names of its
	contributors may be used to endorse or promote products derived
	from this software without specific prior written permission.

THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND
CONTRIBUTORS "AS IS" AND ANY EXPRESS OR IMPLIED WARRANTIES,
INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF
MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR
CONTRIBUTORS BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL,
SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT
LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF
USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED
AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT
LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN
ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE
POSSIBILITY OF SUCH DAMAGE.
"""
import os
import re
import sys
import fitz
from openpyxl import Workbook
from collections import deque
from openpyxl.styles import Font
from table_extract import get_tables
from table_to_str import get_fig_data
from openpyxl.styles import Border, Side
from table_extract import generate_html_files
from get_page_numbers import get_chapter_page_range, get_chapter_number
from argparse import ArgumentParser, ArgumentDefaultsHelpFormatter


def get_text_blocks(file_path, chapter):
    """
    Args:
        file_path : PDF file path
        chapter : chapter number(s) or range of chapters
    """
    doc_obj = fitz.open(file_path)
    load_pages = list(doc_obj.pages())
    start_end_page_info = get_chapter_page_range(file_path, chapter)
    blocks_data = []

    for start_end_page in start_end_page_info:
        for page in load_pages[start_end_page[0] - 1:start_end_page[1]]:
            blocks_data.extend(page.get_text("blocks"))

    return blocks_data


def data_extraction(file_path, out_text_path, out_table_path, projectkey, chapter):
    """
    Args:
        file_path : pdf file path
        out_text_path : Jira upload-able Excel file path
        out_table_path : Excel file path to extract the tables in Non-Jira format
        projectkey : Jira project key
        chapter: Chapter number(s) or range of chapters
    """
    print("\n\n ######--- Text Extraction Started ----#############\n\n")
    j = 0
    l = 0
    all_labels = []
    k = 1
    flag = False
    text_label = False
    work_sheets = []
    label = None
    fig_data = ""
    prev_fig_name = None
    verify_file(file_path)
    pdf_doc_obj = fitz.open(file_path)
    data = get_text_blocks(file_path, chapter)
    par_dir, files = generate_html_files(file_path, out_table_path, chapter)
    fig_name_data = get_fig_data(file_path, par_dir, files)
    thin = Side(border_style="thin")
    wb = Workbook()
    ws = wb.active
    # Get Chapter Names
    chapter_names = get_chapter_names(pdf_doc_obj)

    # Get the Table of Contents (for Sections and subsections)
    toc_data = generate_toc(pdf_doc_obj)
    x = []
    skip_row = 0
    for i in toc_data.values():
        x.extend(i)
    for i in range(len(data)):
        if data[i][5] != 0:
            text = data[i][4]
        else:
            continue
        if "\n" in text:
            text = "".join(text.split("\n"))

        if "<image: DeviceRGB" not in text and re.search(r"\w+", text) and not re.match(r"^[\d]+$", text.strip()) and "http:" not in text and "https:" not in text:
            if text.strip() in x or text.rstrip() in list(toc_data.keys()):
                text_label = True
                label = get_label(text, toc_data)

                if label and label not in all_labels:
                    flag = True
                    all_labels.append(label)
                    j += 1
            else:
                text_label = False

            if label in chapter_names and label not in work_sheets:
                for inv_chr in '[]:*?/\\':
                    if inv_chr in label:
                        label = label.replace(inv_chr, "")
                if len(label) > 31:
                    label = label[:31]

                ws = wb.create_sheet(label)
                ws['A1'].border = Border(bottom=thin, top=thin, left=thin,
                                         right=thin)
                ws['B1'].border = Border(bottom=thin, top=thin, left=thin,
                                         right=thin)
                ws['C1'].border = Border(bottom=thin, top=thin, left=thin,
                                         right=thin)
                ws['D1'].border = Border(bottom=thin, top=thin, left=thin,
                                         right=thin)
                ws['E1'].border = Border(bottom=thin, top=thin, left=thin,
                                         right=thin)
                ws['F1'].border = Border(bottom=thin, top=thin, left=thin,
                                         right=thin)
                ws['G1'].border = Border(bottom=thin, top=thin, left=thin,
                                         right=thin)
                ws.cell(1, 1, "Key")
                ws.cell(1, 2, "Project Key")
                ws.cell(1, 3, "Issue Type")
                ws.cell(1, 4, "Description")
                ws.cell(1, 5, "Summary")
                ws.cell(1, 6, "Labels")
                ws.cell(1, 7, "Is Testable")

                for c in ws.columns:
                    c[0].font = Font(bold=True)

                work_sheets.append(label)
                k = 2

            if flag and not text_label:
                if re.search(r"^Figure \d+:", text.strip()) or re.search(r"^Table \d+ —", text.strip()):
                    fig_name = text.strip()
                    if fig_name in fig_name_data and prev_fig_name != fig_name:
                        fig_data = fig_name_data[fig_name]
                        label = label.replace(" ", "_").replace(",", "_")
                        summary = fig_name + "_" + str(k-1)
                        ws.cell(k, 1, "")
                        ws.cell(k, 2, projectkey)
                        ws.cell(k, 3, "Requirements")
                        ws.cell(k, 4, fig_data)
                        ws.cell(k, 5, summary)
                        ws.cell(k, 6, label)
                        ws.cell(k, 7, "")
                        k += 1
                        skip_row = k
                    prev_fig_name = fig_name
                else:
                    figure_end_start = ""
                    figure_end_last = ""

                    content = text.strip()
                    content_words = content.split()
                    if len(content_words) > 10:
                        content_words_start = content_words[0]
                        content_words_end = " ".join(i for i in content_words[-10:-4])
                    else:
                        content_words_start = content_words[0]
                        content_words_end = content_words[-1]

                    figure_content = [
                        i.replace("||", " ").replace("|", " ").strip() for i in
                        fig_data.splitlines() if re.search(r"\w", i)]
                    if figure_content:
                        figure_end = figure_content[-1].split()
                        figure_end_start = figure_end[0]
                        if len(figure_end) > 10:
                            figure_end_last = " ".join(i for i in figure_end[-10:-4])
                        else:
                            figure_end_last = figure_end[-1]

                    if (content_words_start == figure_end_start) or (
                            content_words_end == figure_end_last):
                        for row in range(skip_row, k):
                            ws.cell(row, 1, "")
                            ws.cell(row, 2, "")
                            ws.cell(row, 3, "")
                            ws.cell(row, 4, "")
                            ws.cell(row, 5, "")
                            ws.cell(row, 6, "")
                            ws.cell(row, 7, "")
                        k = skip_row
                    else:
                        label = label.replace(" ", "_").replace(",", "_")
                        description = label + "_" + str(k - 1)
                        ws.cell(k, 1, "")
                        ws.cell(k, 2, projectkey)
                        ws.cell(k, 3, "Requirements")
                        ws.cell(k, 4, text.strip())
                        ws.cell(k, 5, description)
                        ws.cell(k, 6, label)
                        ws.cell(k, 7, "")
                        k = k + 1

    print("\n %%%%%%%%%%--- Text Extraction Completed and Excel File Generated --- %%%%%%%% \n")
    wb.remove(wb['Sheet'])
    wb.save(out_text_path)
    wb.close()


def get_label(inp, data):
    """
    Args:
        inp: text input
        data: table og content data
    Returns:
        Returns subsection as label
    """
    label = None
    for k, v in data.items():
        if inp.strip() in v:
            if len(v) == 1:
                return inp.strip()
            inp = re.search(r"[\d.]+(.*)", inp).group(1)
        if inp.rstrip() == k:
            if len(v) == 1:
                return v[0]
            label = v[0]
            v = deque(v)
            v.rotate(-1)
            data[k] = list(v)

    return label


def get_chapter_names(pdf_obj):
    """
    To get the chapter names of the pdf file
    Args:
        pdf_obj : Pymupdf file object
    Return:
        Returns the all chapter names from the pdf file
    """
    toc = pdf_obj.get_toc()
    chapter_names = [item[1] for item in toc if item[0] == 1]
    return chapter_names


def generate_toc(pdf_obj):
    """
    Table of contents dictionary
    Args:
        pdf_obj: Pymupdf file object
    """
    section_dict = {}
    for i in pdf_obj.get_toc():
        section_text = re.search(r"[\d.]*(.*)", i[1]).group(1)
        if section_text not in section_dict:
            section_dict[section_text] = [i[1].strip()]
        else:
            section_dict[section_text].append(i[1].strip())
    return section_dict


def verify_file(file_path):
    """
    To verify that the provided file is of pdf type or not
    Args:
        file_path: pdf file path
    """
    file_ext = os.path.splitext(file_path)[1]
    if file_ext.lower() != ".pdf":
        raise Exception("File must be PDF type. Please provide a pdf file")


def verify_python_version():
    """To verify python version"""
    version_info = sys.version_info
    major, minor = version_info.major, version_info.minor

    if major != 3 and minor != 8:
        raise Exception("Please use the python version 3.8")


def get_user_args():
    """To get and parse the command line arguments"""
    cur_work_dir = os.getcwd()
    def_text_out = cur_work_dir + "\data_extract_test.xlsx"
    def_table_out = cur_work_dir + "\data_table_extract_test.xlsx"
    args_parser = ArgumentParser(description='PDF Data Extraction', formatter_class=ArgumentDefaultsHelpFormatter)
    args_parser.add_argument("-f", "--filepath", help="PDF file location", required=True)
    args_parser.add_argument("-o", "--output_filepath", help="output Excel file location", default=def_text_out)
    args_parser.add_argument("-t", "--table_output_filepath", help="output Excel file location for Table", default=def_table_out)
    args_parser.add_argument("-k", "--project_key", help="Project Key of the project", required=True)
    user_input = args_parser.parse_args()
    return user_input


if __name__ == "__main__":
    verify_python_version()
    args = get_user_args()
    pdf_file_path = args.filepath
    excel_path = args.output_filepath
    table_excel_path = args.table_output_filepath
    pr_key = args.project_key
    chapter_number = get_chapter_number(pdf_file_path)
    data_extraction(pdf_file_path, excel_path, table_excel_path, pr_key, chapter_number)
    get_tables(pdf_file_path, table_excel_path, chapter_number)

